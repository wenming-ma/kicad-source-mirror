#!/usr/bin/env python3
"""
Compare two folders and generate an Excel report with statistics and file listings.
"""

import os
import sys
import hashlib
import argparse
from pathlib import Path
from typing import Set, Dict, Tuple, List
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_SKIP_FOLDERS = ['build', 'cmake', 'vcpkg_installed']
DEFAULT_FILE_EXTENSIONS = ['.cpp', '.h', '.hpp', '.txt', '.json']


def get_file_hash(file_path: str) -> str:
    """Calculate MD5 hash of a file."""
    hasher = hashlib.md5()
    try:
        with open(file_path, 'rb') as f:
            for chunk in iter(lambda: f.read(8192), b''):
                hasher.update(chunk)
        return hasher.hexdigest()
    except Exception as e:
        return f"ERROR: {str(e)}"


def should_skip_directory(rel_path: str, skip_folders: List[str], skip_hidden: bool = True) -> bool:
    """Check if a directory should be skipped based on skip list and hidden folder setting."""
    path_parts = rel_path.replace('\\', '/').split('/')

    if skip_hidden:
        for part in path_parts:
            if part.startswith('.') and part != '.':
                return True

    for skip_folder in skip_folders:
        if skip_folder in path_parts:
            return True
    return False


def should_include_file(filename: str, extensions: List[str]) -> bool:
    """Check if a file should be included based on extension list."""
    if not extensions:
        return True
    _, ext = os.path.splitext(filename)
    return ext.lower() in [e.lower() for e in extensions]


def collect_files(folder_path: str, skip_folders: List[str], extensions: List[str], skip_hidden: bool = True) -> Dict[str, str]:
    """
    Collect all files in a folder with their relative paths.
    Returns a dict mapping relative_path -> absolute_path
    """
    files = {}
    folder_path = os.path.abspath(folder_path)

    for root, dirs, filenames in os.walk(folder_path):
        rel_root = os.path.relpath(root, folder_path)

        if should_skip_directory(rel_root, skip_folders, skip_hidden):
            dirs.clear()
            continue

        dirs[:] = [d for d in dirs if not should_skip_directory(
            os.path.relpath(os.path.join(root, d), folder_path), skip_folders, skip_hidden
        )]

        for filename in filenames:
            if not should_include_file(filename, extensions):
                continue

            abs_path = os.path.join(root, filename)
            rel_path = os.path.relpath(abs_path, folder_path)
            rel_path = rel_path.replace('\\', '/')
            files[rel_path] = abs_path

    return files


def compare_folders(folder_a: str, folder_b: str, skip_folders: List[str],
                   extensions: List[str], skip_hidden: bool = True) -> Tuple[Set[str], Set[str], Set[str], Dict[str, bool]]:
    """
    Compare two folders and return:
    - files_only_in_a: Set of relative paths only in folder A
    - files_only_in_b: Set of relative paths only in folder B
    - common_files: Set of relative paths in both folders
    - content_match: Dict mapping relative_path -> True/False (content matches or not)
    """
    files_a = collect_files(folder_a, skip_folders, extensions, skip_hidden)
    files_b = collect_files(folder_b, skip_folders, extensions, skip_hidden)

    set_a = set(files_a.keys())
    set_b = set(files_b.keys())

    files_only_in_a = set_a - set_b
    files_only_in_b = set_b - set_a
    common_files = set_a & set_b

    content_match = {}
    for rel_path in common_files:
        hash_a = get_file_hash(files_a[rel_path])
        hash_b = get_file_hash(files_b[rel_path])
        content_match[rel_path] = (hash_a == hash_b and not hash_a.startswith("ERROR"))

    return files_only_in_a, files_only_in_b, common_files, content_match


def create_excel_report(folder_a: str, folder_b: str, output_file: str,
                       skip_folders: List[str], extensions: List[str], skip_hidden: bool = True):
    """Create Excel report with statistics and file listings."""
    print(f"Comparing folders:")
    print(f"  Folder A: {folder_a}")
    print(f"  Folder B: {folder_b}")
    print(f"  Skip folders: {', '.join(skip_folders)}")
    print(f"  File extensions: {', '.join(extensions) if extensions else 'All'}")
    print(f"  Skip hidden folders (starting with '.'): {'Yes' if skip_hidden else 'No'}")
    print()

    files_only_in_a, files_only_in_b, common_files, content_match = compare_folders(
        folder_a, folder_b, skip_folders, extensions, skip_hidden
    )

    same_content = sum(1 for match in content_match.values() if match)
    diff_content = len(common_files) - same_content

    wb = openpyxl.Workbook()

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    summary_sheet = wb.create_sheet("Statistics", 0)
    only_a_sheet = wb.create_sheet("Only in A", 1)
    only_b_sheet = wb.create_sheet("Only in B", 2)
    common_sheet = wb.create_sheet("Common Files", 3)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    link_font = Font(color="0000FF", underline="single")
    center_align = Alignment(horizontal="center", vertical="center")

    summary_sheet.column_dimensions['A'].width = 40
    summary_sheet.column_dimensions['B'].width = 20

    summary_sheet['A1'] = "Folder Comparison Statistics"
    summary_sheet['A1'].font = Font(bold=True, size=14)
    summary_sheet.merge_cells('A1:B1')

    summary_sheet['A3'] = "Folder A Path:"
    summary_sheet['B3'] = folder_a
    summary_sheet['A4'] = "Folder B Path:"
    summary_sheet['B4'] = folder_b

    summary_sheet['A5'] = "Skip Folders:"
    summary_sheet['B5'] = ', '.join(skip_folders)
    summary_sheet['A6'] = "File Extensions:"
    summary_sheet['B6'] = ', '.join(extensions) if extensions else 'All'
    summary_sheet['A7'] = "Skip Hidden Folders:"
    summary_sheet['B7'] = 'Yes' if skip_hidden else 'No'

    summary_sheet['A9'] = "Category"
    summary_sheet['B9'] = "Count"
    summary_sheet['A9'].fill = header_fill
    summary_sheet['B9'].fill = header_fill
    summary_sheet['A9'].font = header_font
    summary_sheet['B9'].font = header_font
    summary_sheet['A9'].alignment = center_align
    summary_sheet['B9'].alignment = center_align

    row = 10

    summary_sheet[f'A{row}'] = "Total files in Folder A"
    summary_sheet[f'B{row}'] = len(files_only_in_a) + len(common_files)
    row += 1

    summary_sheet[f'A{row}'] = "Total files in Folder B"
    summary_sheet[f'B{row}'] = len(files_only_in_b) + len(common_files)
    row += 1

    summary_sheet[f'A{row}'] = "Files only in Folder A"
    summary_sheet[f'B{row}'] = len(files_only_in_a)
    if len(files_only_in_a) > 0:
        summary_sheet[f'B{row}'].hyperlink = f"#'Only in A'!A1"
        summary_sheet[f'B{row}'].font = link_font
    row += 1

    summary_sheet[f'A{row}'] = "Files only in Folder B"
    summary_sheet[f'B{row}'] = len(files_only_in_b)
    if len(files_only_in_b) > 0:
        summary_sheet[f'B{row}'].hyperlink = f"#'Only in B'!A1"
        summary_sheet[f'B{row}'].font = link_font
    row += 1

    summary_sheet[f'A{row}'] = "Common files (total)"
    summary_sheet[f'B{row}'] = len(common_files)
    if len(common_files) > 0:
        summary_sheet[f'B{row}'].hyperlink = f"#'Common Files'!A1"
        summary_sheet[f'B{row}'].font = link_font
    row += 1

    summary_sheet[f'A{row}'] = "  - Same content"
    summary_sheet[f'B{row}'] = same_content
    row += 1

    summary_sheet[f'A{row}'] = "  - Different content"
    summary_sheet[f'B{row}'] = diff_content
    row += 1

    only_a_sheet.column_dimensions['A'].width = 80
    only_a_sheet['A1'] = "Relative Path"
    only_a_sheet['A1'].fill = header_fill
    only_a_sheet['A1'].font = header_font

    for idx, rel_path in enumerate(sorted(files_only_in_a), start=2):
        only_a_sheet[f'A{idx}'] = rel_path

    only_b_sheet.column_dimensions['A'].width = 80
    only_b_sheet['A1'] = "Relative Path"
    only_b_sheet['A1'].fill = header_fill
    only_b_sheet['A1'].font = header_font

    for idx, rel_path in enumerate(sorted(files_only_in_b), start=2):
        only_b_sheet[f'A{idx}'] = rel_path

    common_sheet.column_dimensions['A'].width = 80
    common_sheet.column_dimensions['B'].width = 20
    common_sheet['A1'] = "Relative Path"
    common_sheet['B1'] = "Content Status"
    common_sheet['A1'].fill = header_fill
    common_sheet['B1'].fill = header_fill
    common_sheet['A1'].font = header_font
    common_sheet['B1'].font = header_font

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for idx, rel_path in enumerate(sorted(common_files), start=2):
        common_sheet[f'A{idx}'] = rel_path
        if content_match[rel_path]:
            common_sheet[f'B{idx}'] = "Same"
            common_sheet[f'B{idx}'].fill = green_fill
        else:
            common_sheet[f'B{idx}'] = "Different"
            common_sheet[f'B{idx}'].fill = red_fill

    wb.save(output_file)

    print(f"Report generated: {output_file}")
    print()
    print("Summary:")
    print(f"  Total files in Folder A: {len(files_only_in_a) + len(common_files)}")
    print(f"  Total files in Folder B: {len(files_only_in_b) + len(common_files)}")
    print(f"  Files only in A: {len(files_only_in_a)}")
    print(f"  Files only in B: {len(files_only_in_b)}")
    print(f"  Common files: {len(common_files)}")
    print(f"    - Same content: {same_content}")
    print(f"    - Different content: {diff_content}")


def main():
    parser = argparse.ArgumentParser(
        description='Compare two folders and generate an Excel report',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  python compare_folders.py folderA folderB output.xlsx
  python compare_folders.py folderA folderB output.xlsx --skip-folders build cmake .git
  python compare_folders.py folderA folderB output.xlsx --extensions .cpp .h .hpp
  python compare_folders.py folderA folderB output.xlsx --skip-folders build --extensions .cpp .h
  python compare_folders.py folderA folderB output.xlsx --include-hidden
        '''
    )

    parser.add_argument('folder_a', help='Path to folder A')
    parser.add_argument('folder_b', help='Path to folder B')
    parser.add_argument('output_file', help='Output Excel file path')
    parser.add_argument(
        '--skip-folders',
        nargs='*',
        default=DEFAULT_SKIP_FOLDERS,
        help=f'Folders to skip during comparison (default: {" ".join(DEFAULT_SKIP_FOLDERS)})'
    )
    parser.add_argument(
        '--extensions',
        nargs='*',
        default=DEFAULT_FILE_EXTENSIONS,
        help=f'File extensions to include (default: {" ".join(DEFAULT_FILE_EXTENSIONS)}). Use --extensions without arguments to include all files.'
    )
    parser.add_argument(
        '--include-hidden',
        action='store_true',
        help='Include hidden folders (folders starting with ".", such as .vs, .vscode). By default, hidden folders are skipped.'
    )

    args = parser.parse_args()

    if not os.path.isdir(args.folder_a):
        print(f"Error: Folder A does not exist: {args.folder_a}")
        sys.exit(1)

    if not os.path.isdir(args.folder_b):
        print(f"Error: Folder B does not exist: {args.folder_b}")
        sys.exit(1)

    create_excel_report(
        args.folder_a,
        args.folder_b,
        args.output_file,
        args.skip_folders,
        args.extensions,
        skip_hidden=not args.include_hidden
    )


if __name__ == "__main__":
    main()