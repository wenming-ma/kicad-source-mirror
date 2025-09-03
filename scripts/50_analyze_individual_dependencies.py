#!/usr/bin/env python3
import json
import os
import subprocess
import sys
import tempfile
from collections import defaultdict
from pathlib import Path

import pandas as pd

# Configuration: File paths and build settings
ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
BUILD_DIR = ROOT / "build" / "x64-Debug"  # Build directory path setting
TU_INDEX_FILE = SCRIPTS_DIR / "tu_index.json"
MINSET_SOURCES_FILE = SCRIPTS_DIR / "pns_sources.json"
RESOLVE_SCRIPT = SCRIPTS_DIR / "30_resolve_minset.py"
OUTPUT_EXCEL = ROOT / "pns_file_dependencies.xlsx"


def load_minset_sources():
    """Load minimum set source files"""
    with open(MINSET_SOURCES_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
        return data.get("sources", [])


def create_temp_seeds_file(target_file):
    """Create temporary seeds.txt file with single target file"""
    temp_seeds = SCRIPTS_DIR / "temp_seeds.txt"
    with open(temp_seeds, "w", encoding="utf-8") as f:
        # Convert absolute path to relative path from KiCad root
        try:
            target_path = Path(target_file)
            # Try to make it relative to ROOT first
            if target_path.is_absolute():
                try:
                    rel_path = target_path.relative_to(ROOT)
                    f.write(str(rel_path).replace("\\", "/") + "\n")
                except ValueError:
                    # If not under ROOT, try common KiCad source paths
                    # Look for pattern like "kicad-source-mirror" and convert
                    path_str = str(target_path)
                    if "kicad-source-mirror" in path_str:
                        # Extract part after kicad-source-mirror
                        idx = path_str.find("kicad-source-mirror")
                        if idx >= 0:
                            remaining = path_str[
                                idx + len("kicad-source-mirror") :
                            ].lstrip("\\/")
                            f.write(remaining.replace("\\", "/") + "\n")
                        else:
                            f.write(target_file + "\n")
                    else:
                        f.write(target_file + "\n")
            else:
                f.write(str(target_path).replace("\\", "/") + "\n")
        except Exception as e:
            print(f"Warning: Could not process path {target_file}: {e}")
            f.write(target_file + "\n")
    return temp_seeds


def run_resolve_minset_for_file(target_file):
    """Run 30_resolve_minset.py for a single file and return dependencies"""
    try:
        # Create temporary seeds file
        temp_seeds = create_temp_seeds_file(target_file)

        # Backup original seeds.txt if it exists
        original_seeds = SCRIPTS_DIR / "seeds.txt"
        backup_seeds = None
        if original_seeds.exists():
            backup_seeds = SCRIPTS_DIR / "seeds_backup.txt"
            # Remove backup if it exists
            if backup_seeds.exists():
                backup_seeds.unlink()
            original_seeds.rename(backup_seeds)

        # Move temp seeds to seeds.txt
        temp_seeds.rename(original_seeds)

        # Run 30_resolve_minset.py
        result = subprocess.run(
            [sys.executable, str(RESOLVE_SCRIPT)],
            cwd=str(ROOT),
            capture_output=True,
            text=True,
            timeout=300,  # 5 minutes timeout
        )

        if result.returncode == 0:
            # Read the generated minset_sources.json
            build_output = BUILD_DIR / "minset_sources.json"
            if build_output.exists():
                with open(build_output, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    dependencies = data.get("sources", [])
            else:
                dependencies = []
        else:
            print(f"Error running resolve_minset for {target_file}: {result.stderr}")
            dependencies = []

        # Restore original seeds.txt
        if backup_seeds and backup_seeds.exists():
            if original_seeds.exists():
                original_seeds.unlink()
            backup_seeds.rename(original_seeds)
        elif original_seeds.exists():
            original_seeds.unlink()

        return dependencies

    except Exception as e:
        print(f"Exception processing {target_file}: {e}")
        return []


def analyze_all_files():
    """Analyze dependencies for all files in minimum set using 30_resolve_minset.py"""
    print("Loading minimum set source files...")

    if not MINSET_SOURCES_FILE.exists():
        print(f"Error: minset_sources.json not found at {MINSET_SOURCES_FILE}")
        return False, {}

    if not RESOLVE_SCRIPT.exists():
        print(f"Error: 30_resolve_minset.py not found at {RESOLVE_SCRIPT}")
        return False, {}

    minset_sources = load_minset_sources()

    print(f"Analyzing {len(minset_sources)} source files...")
    print(f"Each file will be processed using {RESOLVE_SCRIPT}")

    # Results storage
    results = []
    detailed_dependencies = {}

    # Analyze each file by running 30_resolve_minset.py
    for i, source_file in enumerate(minset_sources, 1):
        print(f"Processing {i}/{len(minset_sources)}: {Path(source_file).name}")

        try:
            dependencies = run_resolve_minset_for_file(source_file)

            # Store summary
            results.append(
                {
                    "Source_File": source_file,
                    "File_Name": Path(source_file).name,
                    "Dependency_Count": len(dependencies),
                    "Dependencies": dependencies,
                }
            )

            # Store detailed dependencies for additional sheets
            detailed_dependencies[source_file] = dependencies

        except Exception as e:
            print(f"Error processing {source_file}: {e}")
            results.append(
                {
                    "Source_File": source_file,
                    "File_Name": Path(source_file).name,
                    "Dependency_Count": 0,
                    "Dependencies": [],
                    "Error": str(e),
                }
            )

    return results, detailed_dependencies


def create_excel_report(results, detailed_dependencies):
    """Create Excel report with summary and detail sheets"""
    print(f"Creating Excel report: {OUTPUT_EXCEL}")

    with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
        # Create detail sheets first to get sheet names
        sheet_names_map = {}
        sheet_counter = 1

        for source_file, dependencies in detailed_dependencies.items():
            if sheet_counter > 50:  # Limit number of sheets
                break

            file_name = Path(source_file).stem[:20]  # Limit sheet name length
            sheet_name = f"Detail_{sheet_counter}_{file_name}"
            sheet_names_map[source_file] = sheet_name

            # Create detail dataframe
            detail_df = pd.DataFrame(
                {
                    "Target_File": [source_file],
                    "Total_Dependencies": [len(dependencies)],
                }
            )

            # Add dependencies list
            deps_df = pd.DataFrame(
                {
                    "Dependency_Files": (
                        dependencies if dependencies else ["No dependencies"]
                    )
                }
            )

            # Write to sheet with some spacing
            detail_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
            deps_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=3)

            sheet_counter += 1

        # Create summary sheet with links
        summary_data = []
        for r in results:
            summary_data.append(
                {
                    "Source_File": r["Source_File"],
                    "File_Name": r["File_Name"],
                    "Dependency_Count": r["Dependency_Count"],
                    "View_Details": (
                        f"=HYPERLINK(\"#'{sheet_names_map.get(r['Source_File'], '')}'!A1\", \"View Details\")"
                        if r["Source_File"] in sheet_names_map
                        else "N/A"
                    ),
                }
            )

        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # Format the summary sheet
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows

        workbook = writer.book
        summary_sheet = workbook["Summary"]

        # Style headers
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)

        for cell in summary_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Style the hyperlinks in the View_Details column
        link_font = Font(color="0563C1", underline="single")
        for row in range(2, len(summary_data) + 2):
            link_cell = summary_sheet.cell(row=row, column=4)  # View_Details column
            if link_cell.value and link_cell.value != "N/A":
                link_cell.font = link_font

        # Auto-adjust column widths
        for column in summary_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            summary_sheet.column_dimensions[column_letter].width = adjusted_width

    print(f"Excel report created successfully: {OUTPUT_EXCEL}")


def main():
    """Main function"""
    print("Starting individual file dependency analysis using 30_resolve_minset.py...")
    print("This will run 30_resolve_minset.py for each file in the minimum set")

    # Analyze all files
    results, detailed_dependencies = analyze_all_files()

    if not results:
        print("No results generated. Exiting.")
        return 1

    # Create Excel report
    create_excel_report(results, detailed_dependencies)

    # Print summary statistics
    total_files = len(results)
    avg_dependencies = (
        sum(r["Dependency_Count"] for r in results) / total_files
        if total_files > 0
        else 0
    )
    max_dependencies = max(r["Dependency_Count"] for r in results) if results else 0
    min_dependencies = min(r["Dependency_Count"] for r in results) if results else 0

    print(f"\n=== Analysis Summary ===")
    print(f"Total files analyzed: {total_files}")
    print(f"Average dependencies per file: {avg_dependencies:.2f}")
    print(f"Maximum dependencies: {max_dependencies}")
    print(f"Minimum dependencies: {min_dependencies}")
    print(f"Report saved to: {OUTPUT_EXCEL}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
