#!/usr/bin/env python3
import json
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl for Excel report generation...")
    import subprocess
    import sys

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
INDEX = ROOT / r"scripts\tu_index.json"
SEEDS = ROOT / r"scripts\seeds.txt"
OUT = ROOT / r"scripts\tem\minset_sources.json"
UNRES = ROOT / r"scripts\tem\unresolved_symbols.json"
REPORT = ROOT / r"scripts\tem\dependency_report.xlsx"

# 扩展的外部库符号前缀，避免把系统/三方库当"缺失"
EXTERNAL_HINTS = [
    # C++标准库
    "std::",
    "stdext::",
    "__std",
    "__cxx",
    # 内存管理
    "operator new",
    "operator delete",
    "??_U",
    "??_V",
    "??2",
    "??3",
    # MSVC运行时
    "__imp_",
    "??_",
    "_CRT_",
    "__acrt_",
    "__vcrt_",
    "__security_",
    # Windows API
    "GetProcAddress",
    "LoadLibrary",
    "CreateFile",
    "RegQuery",
    "Nt",
    "Rtl",
    # 线程库
    "pthread_",
    "_beginthreadex",
    "CreateThread",
    "WaitForSingleObject",
    # 系统库
    "dlopen",
    "dlsym",
    "dlclose",
    "__libc_",
    "__glibc_",
    # Boost库
    "boost::",
    "__boost",
    "_boost_",
    # Google库
    "google::",
    "protobuf::",
    "__google_",
    # Qt框架
    "Qt",
    "Q",
    "qt_",
    "__qt_",
    # wxWidgets
    "wx",
    "WX",
    "WXDLLIMPEXP",
    "__wx_",
    # 加密库
    "SSL_",
    "EVP_",
    "CRYPTO_",
    "BN_",
    "RSA_",
    "AES_",
    "SHA1_",
    "SHA256_",
    # 数据库
    "sqlite3_",
    "SQLITE_",
    "mysql_",
    "postgres_",
    # 压缩库
    "zlib",
    "inflate",
    "deflate",
    "compress",
    "uncompress",
    # 图像库
    "png_",
    "jpeg_",
    "tiff_",
    "gif_",
    "webp_",
    # 网络库
    "curl_",
    "SSL_",
    "socket_",
    "inet_",
    # 其他常见第三方库
    "cairo_",
    "pango_",
    "glib_",
    "gtk_",
    "gdk_",
    # OpenGL相关
    "glew",
    "GLEW_",
    "gl",
    "GL_",
    "glm::",
    "glut",
    # 国际化库
    "gettext_",
    "libintl_",
    # 文本塑形库
    "hb_",
    "harfbuzz_",
    # CAD内核
    "OCC",
    "OpenCASCADE",
    "Standard_",
    "Handle_",
    # Unicode库
    "icu_",
    "U_",
    "UCHAR_",
    # 电路仿真
    "ng_",
    "spice_",
    # Git库
    "git_",
    # 消息库
    "nng_",
    # 压缩库扩展
    "ZSTD_",
    "zstd",
    # Python
    "Py",
    "PyObject",
    "_Py",
    # 编译器内建
    "__builtin_",
    "__clang",
    "__GNUC__",
    "_MSC_VER",
]

# 黑名单文件列表 - 定义不想引入的文件
# 当某个符号的定义只在这些文件中时，不引入这些文件，但记录符号信息
BLACKLISTED_FILES = [
    "eeschema/sch_edit_frame.cpp",
    "eeschema/tools/sch_editor_control.cpp",
    "eeschema/tools/symbol_editor_edit_tool.cpp",
    "eeschema/sim/sim_model.cpp",
    "eeschema/sim/sim_lib_mgr.cpp",
    "eeschema/dialogs/dialog_schematic_setup.cpp",
    "common/dialogs/html_message_box.cpp",
    "common/dialog_shim.cpp",
    "eeschema/symbol_editor/symbol_edit_frame.cpp",
    "common/eda_base_frame.cpp",
    "eeschema/tools/sch_selection_tool.cpp",
    "common/tool/selection_tool.cpp",
    "common/tool/tool_manager.cpp",
    "eeschema/sch_base_frame.cpp",
    "eeschema/sch_io/kicad_legacy/sch_io_kicad_legacy.cpp",
    "eeschema/sch_io/sch_io_mgr.cpp",
    "eeschema/sch_commit.cpp",
    "common/dialogs/dialog_migrate_settings.cpp",
    "common/dialogs/dialog_configure_paths.cpp",
    "eeschema/dialogs/dialog_database_lib_settings.cpp",
    "common/dialogs/dialog_hotkey_list.cpp",
    "common/dialog_about/AboutDialog_main.cpp",
    "eeschema/netlist_exporters/netlist_exporter_spice.cpp",
    "eeschema/dialogs/dialog_database_lib_settings_base.cpp",
    "eeschema/sim/spice_settings.cpp",
    "eeschema/sim/spice_value.cpp",
    "common/api/api_plugin_manager.cpp",
    # 可以在这里添加更多不想引入的文件路径
    # 支持相对路径或者绝对路径模式
]


def is_blacklisted_file(file_path):
    """
    检查文件是否在黑名单中
    支持路径模式匹配
    """
    file_path_str = str(file_path)
    for blacklisted in BLACKLISTED_FILES:
        # 检查文件路径是否以黑名单路径结尾
        if file_path_str.endswith(blacklisted) or file_path_str.endswith(
            blacklisted.replace("/", "\\")
        ):
            return True
        # 检查绝对路径匹配
        if file_path_str == blacklisted:
            return True
    return False


def is_external(sym):
    return any(sym.startswith(p) for p in EXTERNAL_HINTS)


def get_seed_files(per, seeds_file):
    """
    Load seed files from seeds.txt and find matching files in the index

    Args:
        per: Dictionary mapping file paths to their symbols
        seeds_file: Path to the seeds.txt file

    Returns:
        Set of matching absolute file paths
    """
    # Load seed paths and convert to Path objects
    seeds = []
    for s in seeds_file.read_text(encoding="utf-8").splitlines():
        if s.strip() and not s.strip().startswith("#"):
            seeds.append(Path(s.strip()))

    # Find matching seeds using Path suffix matching
    S = set()
    for seed in seeds:
        for abs_path in per.keys():
            # Check if the absolute path ends with the seed path
            if abs_path.match(f"*/{seed}") or abs_path.match(f"*\\{seed}"):
                S.add(abs_path)
                break
            # Also try direct suffix check
            try:
                if abs_path.relative_to(abs_path.anchor).match(
                    str(seed).replace("\\", "/")
                ):
                    S.add(abs_path)
                    break
            except ValueError:
                pass
            # Fallback: check if parts match
            if len(seed.parts) <= len(abs_path.parts):
                if abs_path.parts[-len(seed.parts) :] == seed.parts:
                    S.add(abs_path)
                    break

    return S


def track_symbol_dependencies(S, per):
    """
    Track which files require which symbols
    Returns: dict mapping symbol -> set of files that require it
    """
    symbol_requirers = defaultdict(set)
    for file in S:
        _, undefined = per.get(file, (set(), set()))
        for sym in undefined:
            symbol_requirers[sym].add(file)
    return symbol_requirers


def create_excel_report(iteration_reports, blacklisted_symbols=None):
    """
    Create Excel report with dependency tracking
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    link_font = Font(color="0563C1", underline="single")

    # Create summary sheet
    summary_ws = wb.create_sheet("Summary", 0)
    summary_ws.append(
        ["Iteration", "Initial Files", "Added Files", "Final Files", "Missing Symbols"]
    )
    for cell in summary_ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # First pass: Create all iteration sheets at the front
    detail_sheets_to_create = []  # Store detail sheet info for later creation

    for iter_num, report in enumerate(iteration_reports, 1):
        # Create iteration sheet
        iter_sheet_name = f"Iteration_{iter_num}"
        iter_ws = wb.create_sheet(iter_sheet_name)

        # Headers
        iter_ws.append(["File", "Introduced Files Count", "Missing Symbols Resolved"])
        for cell in iter_ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Data rows
        row_num = 2
        for file_path, info in sorted(report["file_introductions"].items()):
            # Get relative path from ROOT directory
            try:
                rel_path = Path(file_path).relative_to(ROOT)
                file_display = str(rel_path)
            except ValueError:
                # If can't get relative path, use the filename
                file_display = Path(file_path).name

            introduced_count = len(info["introduced_files"])
            symbols_count = len(info["symbols_resolved"])

            iter_ws.cell(row=row_num, column=1, value=file_display)

            if introduced_count > 0:
                # Store detail sheet info for later creation
                detail_sheet_name = f"Detail_{iter_num}_{row_num}"
                detail_sheets_to_create.append(
                    {
                        "sheet_name": detail_sheet_name,
                        "file_name": file_display,
                        "iter_num": iter_num,
                        "iter_sheet_name": iter_sheet_name,
                        "row_num": row_num,
                        "introduced_files": info["introduced_files"],
                    }
                )

                # Create hyperlink in count cell
                count_cell = iter_ws.cell(row=row_num, column=2, value=introduced_count)
                count_cell.hyperlink = f"#'{detail_sheet_name}'!A1"
                count_cell.font = link_font
            else:
                iter_ws.cell(row=row_num, column=2, value=0)

            iter_ws.cell(row=row_num, column=3, value=symbols_count)
            row_num += 1

        # Auto-adjust column widths for iteration sheet
        for column in iter_ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            iter_ws.column_dimensions[get_column_letter(column[0].column)].width = (
                adjusted_width
            )

        # Add summary row
        summary_ws.append(
            [
                iter_num,
                report["initial_files"],
                report["added_files"],
                report["final_files"],
                report["missing_symbols"],
            ]
        )

    # Second pass: Create all detail sheets at the end
    for detail_info in detail_sheets_to_create:
        detail_ws = wb.create_sheet(detail_info["sheet_name"])
        detail_ws.append(
            [
                f"Files introduced by {detail_info['file_name']} in Iteration {detail_info['iter_num']}"
            ]
        )
        detail_ws["A1"].font = Font(bold=True, size=12)
        detail_ws.merge_cells("A1:C1")

        detail_ws.append(["Introduced File", "Symbol", "Type"])
        for cell in detail_ws[2]:
            cell.font = header_font
            cell.fill = header_fill

        for introduced_file, symbol in detail_info["introduced_files"]:
            # Get relative path for introduced file
            try:
                rel_path = Path(introduced_file).relative_to(ROOT)
                file_display = str(rel_path)
            except ValueError:
                # If can't get relative path, use the filename
                file_display = Path(introduced_file).name

            detail_ws.append(
                [
                    file_display,
                    symbol,
                    "Required Symbol",
                ]
            )

        # Add back link
        detail_ws.append([])
        back_cell = detail_ws.cell(
            row=detail_ws.max_row + 1, column=1, value="<< Back to iteration sheet"
        )
        back_cell.hyperlink = (
            f"#'{detail_info['iter_sheet_name']}'!A{detail_info['row_num']}"
        )
        back_cell.font = link_font

        # Auto-adjust column widths
        for column in detail_ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 100)
            detail_ws.column_dimensions[get_column_letter(column[0].column)].width = (
                adjusted_width
            )

    # Add blacklisted symbols section if available
    if blacklisted_symbols:
        # Add some spacing
        summary_ws.append([])
        summary_ws.append([])

        # Add blacklisted symbols header
        summary_ws.append(["Blacklisted Symbols", "", "", "", ""])
        blacklist_header_row = summary_ws.max_row
        for col in range(1, 6):
            cell = summary_ws.cell(row=blacklist_header_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
        summary_ws.merge_cells(f"A{blacklist_header_row}:E{blacklist_header_row}")

        # Add blacklisted symbols table header
        summary_ws.append(["Symbol", "Blacklisted Files", "", "", ""])
        blacklist_table_header_row = summary_ws.max_row
        for col in range(1, 3):
            cell = summary_ws.cell(row=blacklist_table_header_row, column=col)
            cell.font = header_font
            cell.fill = header_fill

        # Add blacklisted symbols data
        for symbol, files in sorted(blacklisted_symbols.items()):
            # Get relative paths from the root directory
            file_paths = []
            for f in files:
                try:
                    # Try to get relative path from ROOT
                    rel_path = Path(f).relative_to(ROOT)
                    file_paths.append(str(rel_path))
                except ValueError:
                    # If can't get relative path, use the filename
                    file_paths.append(Path(f).name)

            file_list = (
                ", ".join(file_paths)
                if len(file_paths) <= 2
                else f"{', '.join(file_paths[:2])} ... ({len(file_paths)} files)"
            )
            summary_ws.append(
                [symbol, file_list, "", "", ""]
            )

    # Auto-adjust summary sheet columns
    for column in summary_ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(
            max_length + 2, 50
        )  # Increased max width for blacklist info
        summary_ws.column_dimensions[get_column_letter(column[0].column)].width = (
            adjusted_width
        )

    return wb


def main():
    idx_all = json.loads(INDEX.read_text(encoding="utf-8"))["items"]

    # Build mapping using Path suffix matching
    per = {}
    providers = {}

    for it in idx_all:
        abs_path = Path(it["src"])
        per[abs_path] = (set(it["defined"]), set(it["undefined"]))

        for s in it["defined"]:
            providers.setdefault(s, set()).add(abs_path)

    # Get seed files
    S = get_seed_files(per, SEEDS)
    print(f"Initial seed files: {len(S)} files")

    # Track reports for each iteration
    iteration_reports = []

    # Track symbols blocked by blacklisted files
    blacklisted_symbols = {}  # symbol -> set of blacklisted files that provide it

    iteration = 0
    changed = True
    while changed:
        iteration += 1
        changed = False

        # Record initial state
        S_before = S.copy()
        initial_file_count = len(S)

        # Track which files require which symbols
        symbol_requirers = track_symbol_dependencies(S, per)

        provided = set()
        required = set()

        # 收集当前集合的符号信息
        for s in S:
            d, u = per.get(s, (set(), set()))
            provided |= d
            required |= u

        # 找出缺失的内部符号
        missing = {m for m in (required - provided) if not is_external(m)}
        initial_missing_count = len(missing)

        print(
            f"Iteration {iteration}: size={len(S)}, provided={len(provided)}, required={len(required)}, missing={initial_missing_count}"
        )

        # Track introductions for this iteration
        file_introductions = defaultdict(
            lambda: {"introduced_files": [], "symbols_resolved": set()}
        )

        # 为缺失符号找提供者
        added_this_round = 0
        for m in sorted(missing):
            # 检查S中是否已有文件提供了符号m
            # 注意：S在循环中会更新，所以需要实时检查
            providers_for_m = providers.get(m, set())
            if not providers_for_m.intersection(S):
                # S中没有文件提供m，需要添加一个提供者
                if providers_for_m:
                    # 分离黑名单和非黑名单的提供者
                    non_blacklisted_providers = {
                        p for p in providers_for_m if not is_blacklisted_file(p)
                    }
                    blacklisted_providers = {
                        p for p in providers_for_m if is_blacklisted_file(p)
                    }

                    if non_blacklisted_providers:
                        # 如果有非黑名单的提供者，选择字典序最小的
                        chosen = sorted(non_blacklisted_providers)[0]

                        # Track which files required this symbol
                        requiring_files = symbol_requirers.get(m, set())
                        for req_file in requiring_files:
                            file_introductions[str(req_file)][
                                "introduced_files"
                            ].append((str(chosen), m))
                            file_introductions[str(req_file)]["symbols_resolved"].add(m)

                        S.add(chosen)
                        added_this_round += 1
                        changed = True
                    else:
                        # 所有提供者都在黑名单中，记录这个符号但不添加文件
                        blacklisted_symbols[m] = blacklisted_providers
                        print(
                            f"  Symbol '' blocked by blacklisted files: {[str(p) for p in blacklisted_providers]}"
                        )
            # 否则跳过m，因为S中已有文件提供了它

        if added_this_round > 0:
            print(f"  -> Added {added_this_round} new source files")

        # Record iteration report
        iteration_report = {
            "iteration": iteration,
            "initial_files": initial_file_count,
            "added_files": added_this_round,
            "final_files": len(S),
            "missing_symbols": initial_missing_count,
            "file_introductions": dict(file_introductions),
        }

        # Fill in empty entries for files that didn't introduce anything
        for file in S_before:
            file_str = str(file)
            if file_str not in iteration_report["file_introductions"]:
                iteration_report["file_introductions"][file_str] = {
                    "introduced_files": [],
                    "symbols_resolved": set(),
                }

        iteration_reports.append(iteration_report)

        # 防止无限循环
        if iteration > 100:
            print("Warning: Maximum iterations reached, possible circular dependency")
            break

    print(f"Symbol closure converged, final set size: {len(S)} source files")

    provided = set()
    required = set()
    for s in S:
        d, u = per.get(s, (set(), set()))
        provided |= d
        required |= u
    unresolved = sorted(required - provided)

    # Ensure output directory exists
    OUT.parent.mkdir(parents=True, exist_ok=True)

    OUT.write_text(
        json.dumps({"sources": sorted(str(s) for s in S)}, indent=2), encoding="utf-8"
    )
    UNRES.write_text(
        json.dumps({"unresolved_symbols": unresolved}, indent=2), encoding="utf-8"
    )
    print(f"Wrote {OUT}; unresolved={len(unresolved)}")

    # Generate Excel report
    if iteration_reports:
        print(f"\nGenerating Excel dependency report...")
        wb = create_excel_report(iteration_reports, blacklisted_symbols)
        wb.save(REPORT)
        print(f"Wrote dependency report to {REPORT}")
        print(f"  - Summary sheet with {len(iteration_reports)} iterations")
        print(f"  - Detailed sheets for files that introduced new dependencies")
        print(f"  - Click on introduction counts to see detailed file lists")
        if blacklisted_symbols:
            print(
                f"  - Blacklisted symbols section with {len(blacklisted_symbols)} symbols blocked by blacklisted files"
            )


if __name__ == "__main__":
    main()
