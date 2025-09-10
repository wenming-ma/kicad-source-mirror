#!/usr/bin/env python3
"""
Verify and display structure of the generated Excel dependency report
"""
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
REPORT = ROOT / r"scripts\tem\dependency_report.xlsx"

def verify_report():
    """Verify and display Excel report structure"""
    if not REPORT.exists():
        print(f"Report file not found: {REPORT}")
        return
    
    print(f"Loading Excel report from: {REPORT}")
    print(f"File size: {REPORT.stat().st_size:,} bytes\n")
    
    wb = load_workbook(REPORT, read_only=True)
    
    print("=" * 60)
    print("EXCEL REPORT STRUCTURE")
    print("=" * 60)
    
    # Summary sheet
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        print("\n[Summary Sheet]")
        print(f"  Rows: {ws.max_row}, Columns: {ws.max_column}")
        
        # Read header and first few data rows
        headers = []
        for col in range(1, ws.max_column + 1):
            headers.append(ws.cell(1, col).value)
        print(f"  Headers: {headers}")
        
        # Show iteration summary
        print("\n  Iteration Summary:")
        for row in range(2, min(5, ws.max_row + 1)):
            values = [ws.cell(row, col).value for col in range(1, 6)]
            print(f"    Iter {values[0]}: {values[1]} -> {values[3]} files (+{values[2]}), {values[4]} missing symbols")
    
    # Iteration sheets
    iteration_sheets = [s for s in wb.sheetnames if s.startswith("Iteration_")]
    print(f"\n[Iteration Sheets]: {len(iteration_sheets)} sheets")
    
    for sheet_name in iteration_sheets[:2]:  # Show first 2 iterations
        ws = wb[sheet_name]
        print(f"\n  {sheet_name}:")
        print(f"    Rows: {ws.max_row}, Columns: {ws.max_column}")
        
        # Count files with introductions
        files_with_intros = 0
        total_intros = 0
        for row in range(2, ws.max_row + 1):
            count = ws.cell(row, 2).value
            if count and count > 0:
                files_with_intros += 1
                total_intros += count
        
        print(f"    Files that introduced new files: {files_with_intros}")
        print(f"    Total new files introduced: {total_intros}")
        
        # Show top introducers
        introducers = []
        for row in range(2, min(ws.max_row + 1, 10)):
            file_name = ws.cell(row, 1).value
            count = ws.cell(row, 2).value
            if count and count > 0:
                introducers.append((file_name, count))
        
        if introducers:
            print(f"    Top introducers:")
            for name, count in sorted(introducers, key=lambda x: x[1], reverse=True)[:3]:
                print(f"      - {name}: {count} files")
    
    # Detail sheets
    detail_sheets = [s for s in wb.sheetnames if s.startswith("Detail_")]
    print(f"\n[Detail Sheets]: {len(detail_sheets)} sheets")
    print("  These sheets contain specific file introductions with symbol details")
    
    if detail_sheets:
        # Show sample detail sheet
        ws = wb[detail_sheets[0]]
        print(f"\n  Sample: {detail_sheets[0]}")
        print(f"    Rows: {ws.max_row}, Columns: {ws.max_column}")
        
        # Count unique files and symbols
        files = set()
        symbols = set()
        for row in range(3, min(ws.max_row, 10)):  # Sample first few rows
            file_val = ws.cell(row, 1).value
            sym_val = ws.cell(row, 2).value
            if file_val:
                files.add(file_val)
            if sym_val:
                symbols.add(sym_val[:50])  # Truncate for display
        
        print(f"    Sample introduced files: {len(files)}")
        print(f"    Sample symbols resolved: {len(symbols)}")
    
    print("\n" + "=" * 60)
    print("REPORT FEATURES:")
    print("=" * 60)
    print("1. Summary sheet with iteration overview")
    print("2. One sheet per iteration showing file introductions")
    print("3. Clickable links on introduction counts")
    print("4. Detail sheets for each file's specific introductions")
    print("5. Symbol-level tracking for dependency analysis")
    print("6. Back navigation links in detail sheets")
    
    wb.close()
    print("\nReport verification complete!")

if __name__ == "__main__":
    verify_report()