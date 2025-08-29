#!/usr/bin/env python3
"""
批量依赖分析脚本 - 分析每个cpp文件的独立依赖并生成Excel报告

使用方法:
    python batch_dependency_analysis.py all_dependencies.txt

功能:
1. 读取all_dependencies.txt中的所有cpp文件
2. 对每个cpp文件单独运行transitive_closure.py
3. 生成Excel报告，包含:
   - 首页: 各cpp文件的依赖数量汇总和跳转链接
   - 各Sheet页: 每个cpp文件的详细依赖列表
"""

import os
import sys
import subprocess
import tempfile
import time
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def read_cpp_files(dependencies_file):
    """从all_dependencies.txt中提取所有cpp文件"""
    cpp_files = []
    try:
        with open(dependencies_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line.endswith('.cpp'):
                    cpp_files.append(line)
    except Exception as e:
        print(f"错误: 读取文件 {dependencies_file} 失败: {e}")
        sys.exit(1)
    
    print(f"找到 {len(cpp_files)} 个cpp文件")
    return cpp_files

def run_transitive_closure(cpp_file):
    """对单个cpp文件运行transitive_closure.py"""
    try:
        # 创建临时种子文件
        with tempfile.NamedTemporaryFile(mode='w', suffix='_seeds.txt', delete=False) as temp_seeds:
            temp_seeds.write(cpp_file + '\n')
            temp_seeds_path = temp_seeds.name
        
        # 创建临时输出文件
        temp_output = tempfile.NamedTemporaryFile(mode='w', suffix='_deps.txt', delete=False)
        temp_output_path = temp_output.name
        temp_output.close()
        
        # 为每个文件创建唯一的输出文件名，避免覆盖原始all_dependencies.txt
        safe_filename = cpp_file.replace('/', '_').replace('\\', '_').replace(':', '_')
        temp_output_file = f"temp_deps_{safe_filename}.txt"
        
        # 运行transitive_closure.py
        cmd = [
            'python', 
            'transitive_closure.py',
            cpp_file,  # 直接作为命令行参数传递
            '--output', temp_output_file  # 指定临时输出文件
        ]
        
        print(f"  分析: {cpp_file}")
        
        # 运行脚本
        process = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            cwd=os.getcwd()
        )
        
        # 等待脚本完成
        stdout, stderr = process.communicate()
        
        if process.returncode != 0:
            print(f"    警告: 分析 {cpp_file} 失败")
            print(f"    stderr: {stderr}")
            return []
        
        # 解析输出获取依赖文件
        dependencies = []
        
        # 读取临时输出文件
        if os.path.exists(temp_output_file):
            try:
                with open(temp_output_file, 'r', encoding='utf-8') as f:
                    dependencies = [line.strip() for line in f if line.strip()]
            except Exception as e:
                print(f"    错误: 读取输出文件失败: {e}")
        else:
            print(f"    警告: 输出文件 {temp_output_file} 不存在")
        
        # 清理临时文件
        try:
            os.unlink(temp_seeds_path)
            os.unlink(temp_output_path)
            if os.path.exists(temp_output_file):
                os.unlink(temp_output_file)  # 清理临时依赖文件
        except Exception as e:
            pass
            
        print(f"    找到 {len(dependencies)} 个依赖文件")
        return dependencies
        
    except Exception as e:
        print(f"    错误: 分析 {cpp_file} 时发生异常: {e}")
        return []

def create_excel_report(analysis_results, output_file):
    """创建Excel报告"""
    print(f"\n生成Excel报告: {output_file}")
    
    wb = Workbook()
    
    # 删除默认工作表
    wb.remove(wb.active)
    
    # 创建汇总页
    summary_sheet = wb.create_sheet(title="汇总")
    
    # 设置标题
    summary_sheet['A1'] = 'cpp文件依赖分析汇总报告'
    summary_sheet['A1'].font = Font(size=16, bold=True)
    summary_sheet['A1'].alignment = Alignment(horizontal='center')
    summary_sheet.merge_cells('A1:D1')
    
    # 设置表头
    headers = ['序号', 'cpp文件', '依赖文件数量', '详细信息']
    for col, header in enumerate(headers, 1):
        cell = summary_sheet.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 填充数据并创建各个详细页面
    total_files = len(analysis_results)
    total_deps = 0
    
    for idx, (cpp_file, dependencies) in enumerate(analysis_results.items(), 1):
        # 汇总页数据
        summary_sheet.cell(row=idx+3, column=1, value=idx)
        summary_sheet.cell(row=idx+3, column=2, value=cpp_file)
        summary_sheet.cell(row=idx+3, column=3, value=len(dependencies))
        
        # 创建跳转链接
        link_cell = summary_sheet.cell(row=idx+3, column=4, value="查看详情")
        sheet_name = f"Sheet{idx}"
        link_cell.hyperlink = f"#{sheet_name}!A1"
        link_cell.font = Font(color="0000FF", underline="single")
        
        # 创建详细信息页面
        detail_sheet = wb.create_sheet(title=sheet_name)
        
        # 详细页面标题
        detail_sheet['A1'] = f'{cpp_file} - 依赖分析'
        detail_sheet['A1'].font = Font(size=14, bold=True)
        detail_sheet.merge_cells('A1:B1')
        
        # 返回汇总页链接
        back_link = detail_sheet['A2']
        back_link.value = "← 返回汇总"
        back_link.hyperlink = "#汇总!A1"
        back_link.font = Font(color="0000FF", underline="single")
        
        # 依赖文件列表标题
        detail_sheet['A4'] = '序号'
        detail_sheet['B4'] = '依赖文件'
        detail_sheet['A4'].font = Font(bold=True)
        detail_sheet['B4'].font = Font(bold=True)
        
        # 填充依赖文件
        for dep_idx, dep_file in enumerate(dependencies, 1):
            detail_sheet.cell(row=dep_idx+4, column=1, value=dep_idx)
            detail_sheet.cell(row=dep_idx+4, column=2, value=dep_file)
        
        # 调整列宽
        detail_sheet.column_dimensions['A'].width = 8
        detail_sheet.column_dimensions['B'].width = 80
        
        total_deps += len(dependencies)
    
    # 添加汇总统计
    summary_row = total_files + 5
    summary_sheet.cell(row=summary_row, column=1, value="总计:")
    summary_sheet.cell(row=summary_row, column=1).font = Font(bold=True)
    summary_sheet.cell(row=summary_row, column=2, value=f"{total_files} 个cpp文件")
    summary_sheet.cell(row=summary_row, column=3, value=f"{total_deps} 个依赖")
    
    # 调整汇总页列宽
    summary_sheet.column_dimensions['A'].width = 8
    summary_sheet.column_dimensions['B'].width = 50
    summary_sheet.column_dimensions['C'].width = 15
    summary_sheet.column_dimensions['D'].width = 15
    
    # 保存文件
    wb.save(output_file)
    print(f"Excel报告已保存: {output_file}")
    print(f"包含 {total_files} 个cpp文件的分析结果，共 {total_deps} 个依赖关系")

def main():
    if len(sys.argv) != 2:
        print("使用方法: python batch_dependency_analysis.py all_dependencies.txt")
        sys.exit(1)
    
    dependencies_file = sys.argv[1]
    
    if not os.path.exists(dependencies_file):
        print(f"错误: 文件 {dependencies_file} 不存在")
        sys.exit(1)
    
    print("=== KiCad依赖分析批处理脚本 ===")
    print(f"输入文件: {dependencies_file}")
    
    # 读取cpp文件列表
    cpp_files = read_cpp_files(dependencies_file)
    
    if not cpp_files:
        print("错误: 没有找到cpp文件")
        sys.exit(1)
    
    # 分析每个cpp文件
    print(f"\n开始分析 {len(cpp_files)} 个cpp文件...")
    analysis_results = {}
    
    start_time = time.time()
    
    for i, cpp_file in enumerate(cpp_files, 1):
        print(f"[{i}/{len(cpp_files)}] 处理: {cpp_file}")
        dependencies = run_transitive_closure(cpp_file)
        analysis_results[cpp_file] = dependencies
        
        # 每10个文件显示一次进度
        if i % 10 == 0:
            elapsed = time.time() - start_time
            estimated_total = elapsed * len(cpp_files) / i
            remaining = estimated_total - elapsed
            print(f"  进度: {i}/{len(cpp_files)} ({i/len(cpp_files)*100:.1f}%), 预计剩余: {remaining/60:.1f}分钟")
    
    elapsed_time = time.time() - start_time
    print(f"\n分析完成，用时: {elapsed_time/60:.1f}分钟")
    
    # 生成Excel报告
    output_file = "cpp_dependencies_analysis.xlsx"
    create_excel_report(analysis_results, output_file)
    
    print(f"\n=== 分析完成 ===")
    print(f"总计分析了 {len(cpp_files)} 个cpp文件")
    print(f"Excel报告: {output_file}")

if __name__ == "__main__":
    main()